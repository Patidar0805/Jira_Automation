"""
Jira Status Tracker → Excel (Changelog Edition)
- Tracks ALL activity on tickets while assigned to YOU
- Each status change and comment = separate row in Excel
- Uses Jira Changelog API to capture every individual change
- Runs daily at 8 PM via cron, but logs everything that happened during the day
"""

import requests
import json
import os
from datetime import datetime, timezone, timedelta
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIG — Fill in your details here
# ─────────────────────────────────────────────
JIRA_CONFIG = {
    "base_url":  "",   # e.g. https://mycompany.atlassian.net
    "email":     "",               # Your Jira login email
    "api_token": "",                  # https://id.atlassian.com/manage-profile/security/api-tokens
    "projects":   [],                            
# ─────────────────────────────────────────────
#  JIRA API
# ─────────────────────────────────────────────

OUTPUT_FILE = "jira_status_log.xlsx"
STATE_FILE  = "jira_state.json"

# Only these 2 status transitions will be logged
TRACKED_TRANSITIONS = [
    ("To-Do",          "In Development"),
    ("To Do",          "In Development"),
    ("In Development", "Code Review"),
]

# Soft colors cycling per ticket group
TICKET_GROUP_COLORS = [
    "E3F2FD", "F3E5F5", "E8F5E9", "FFF8E1",
    "FCE4EC", "E0F7FA", "FBE9E7", "EDE7F6",
    "F1F8E9", "E8EAF6",
]

# ─────────────────────────────────────────────
#  JIRA API
# ─────────────────────────────────────────────
def auth():
    return HTTPBasicAuth(JIRA_CONFIG["email"], JIRA_CONFIG["api_token"])

def get_my_account_id():
    resp = requests.get(f"{JIRA_CONFIG['base_url']}/rest/api/3/myself", auth=auth())
    resp.raise_for_status()
    return resp.json()["accountId"]

def fetch_my_issues(account_id):
    projects_jql   = ", ".join(JIRA_CONFIG["projects"])
    first_of_month = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    jql = (
        f"project in ({projects_jql}) AND ("
        f"assignee = \"{account_id}\" OR "
        f"(assignee was \"{account_id}\" AND updatedDate >= \"{first_of_month}\")"
        f") ORDER BY updated DESC"
    )
    resp = requests.get(
        f"{JIRA_CONFIG['base_url']}/rest/api/3/search/jql",
        auth=auth(),
        params={"jql": jql, "fields": "summary,status,assignee,priority,comment", "maxResults": 200}
    )
    resp.raise_for_status()
    return resp.json().get("issues", [])

def fetch_changelog(issue_key):
    url, all_entries, start = f"{JIRA_CONFIG['base_url']}/rest/api/3/issue/{issue_key}/changelog", [], 0
    while True:
        resp = requests.get(url, auth=auth(), params={"startAt": start, "maxResults": 100})
        resp.raise_for_status()
        data   = resp.json()
        values = data.get("values", [])
        all_entries.extend(values)
        if start + len(values) >= data.get("total", 0):
            break
        start += len(values)
    return all_entries

def fetch_comments(issue_key):
    resp = requests.get(f"{JIRA_CONFIG['base_url']}/rest/api/3/issue/{issue_key}/comment", auth=auth(), params={"maxResults": 200})
    resp.raise_for_status()
    return resp.json().get("comments", [])

def extract_comment_text(body):
    if isinstance(body, str):
        return body
    return " ".join(
        inline.get("text", "")
        for block in body.get("content", [])
        for inline in block.get("content", [])
        if inline.get("type") == "text"
    ).strip()

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def get_my_assignment_windows(changelog, my_account_id):
    windows, assigned_at = [], None
    for entry in sorted(changelog, key=lambda x: x["created"]):
        for item in entry.get("items", []):
            if item["field"] == "assignee":
                if item.get("to") == my_account_id:
                    assigned_at = entry["created"]
                elif item.get("from") == my_account_id and assigned_at:
                    windows.append((assigned_at, entry["created"]))
                    assigned_at = None
    if assigned_at:
        windows.append((assigned_at, None))
    return windows

def was_assigned_to_me(timestamp, windows):
    return any(
        (end is None and timestamp >= start) or (end and start <= timestamp <= end)
        for start, end in windows
    )

def is_current_month(timestamp):
    ist = datetime.strptime(timestamp[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc) + timedelta(hours=11, minutes=30)
    now = datetime.now()
    return ist.year == now.year and ist.month == now.month

def to_ist(timestamp):
    utc = datetime.strptime(timestamp[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)
    return (utc + timedelta(hours=11, minutes=30)).strftime("%Y-%m-%d %H:%M:%S")

def is_tracked_transition(old_status, new_status):
    return (old_status, new_status) in TRACKED_TRANSITIONS

# ─────────────────────────────────────────────
#  STATE
# ─────────────────────────────────────────────
def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            return json.load(f)
    return {}

def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)

# ─────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────
# ★ Done At is now BEFORE Updated By
HEADERS    = ["Ticket", "Summary", "Activity", "Old Status", "New Status", "Comment", "Done At", "Updated By", "Priority"]
COL_WIDTHS = [14,       38,        22,         18,           18,           45,        22,         22,           12]

def get_month_sheet_name():
    return datetime.now().strftime("%B-%Y")

def style_header_row(ws):
    hfill = PatternFill("solid", start_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

def get_or_create_month_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    style_header_row(ws)
    return ws

def load_or_create_workbook():
    if os.path.exists(OUTPUT_FILE):
        return load_workbook(OUTPUT_FILE)
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def write_data_row(ws, row_num, row_data, bg_color):
    row_fill = PatternFill("solid", start_color=bg_color)
    border   = Border(bottom=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"))
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font      = Font(name="Arial", size=10, bold=(col == 1))
        cell.fill      = row_fill
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
        cell.border    = border
    # Auto row height based on comment
    comment_text = str(row_data[5] or "")
    line_count   = max(1, comment_text.count("\n") + 1, len(comment_text) // 60 + 1)
    ws.row_dimensions[row_num].height = max(20, line_count * 15)

def write_blank_row(ws, row_num):
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col, value=None)
        cell.fill = PatternFill("solid", start_color="FFFFFF")
    ws.row_dimensions[row_num].height = 8

def read_existing_sheet_data(ws):
    ticket_rows_map = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            key = str(row[0] or "").strip()
            if key:
                ticket_rows_map[key].append(list(row))
    return ticket_rows_map

def rebuild_sheet(ws, ticket_rows_map):
    """Rebuild entire sheet: groups sorted by earliest date desc, rows within group asc."""
    # Clear sheet
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            cell.fill  = PatternFill("solid", start_color="FFFFFF")

    if not ticket_rows_map:
        return

    # Sort rows within each ticket ascending by Done At (col index 6)
    for key in ticket_rows_map:
        ticket_rows_map[key].sort(key=lambda x: str(x[6] or ""))

    # Sort groups by earliest date descending
    sorted_tickets = sorted(
        ticket_rows_map.items(),
        key=lambda x: str(x[1][0][6] or ""),
        reverse=True
    )

    current_row = 2
    for color_index, (_, rows_sorted) in enumerate(sorted_tickets):
        bg_color = TICKET_GROUP_COLORS[color_index % len(TICKET_GROUP_COLORS)]
        for row_data in rows_sorted:
            write_data_row(ws, current_row, row_data, bg_color)
            current_row += 1
        write_blank_row(ws, current_row)
        current_row += 1

def update_run_log(wb, total_logged, log_entries, total_tickets):
    log_sheet_name = "Run Logs"
    if log_sheet_name in wb.sheetnames:
        wl = wb[log_sheet_name]
    else:
        wl = wb.create_sheet(title=log_sheet_name)
        log_headers = ["Run Date", "Run Time", "Tickets Found", "Total Changes", "Activity Log"]
        hfill = PatternFill("solid", start_color="1F4E79")
        hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        for i, (h, w) in enumerate(zip(log_headers, [16, 12, 16, 16, 80]), start=1):
            cell = wl.cell(row=1, column=i, value=h)
            cell.font      = hfont
            cell.fill      = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            wl.column_dimensions[get_column_letter(i)].width = w
        wl.row_dimensions[1].height = 24
        wl.freeze_panes = "A2"

    now      = datetime.now()
    next_row = wl.max_row + 1
    bg_color = "F5F5F5" if next_row % 2 == 0 else "FFFFFF"
    row_fill = PatternFill("solid", start_color=bg_color)
    border   = Border(bottom=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"))

    row_data = [
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        total_tickets,
        total_logged,
        "\n".join(log_entries) if log_entries else "No new activity found."
    ]
    for col, value in enumerate(row_data, start=1):
        cell = wl.cell(row=next_row, column=col, value=value)
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col == 5))
        cell.border    = border
    wl.row_dimensions[next_row].height = max(20, len(log_entries) * 15 if log_entries else 20)

# ─────────────────────────────────────────────
#  MAIN SYNC
# ─────────────────────────────────────────────
def sync():
    now        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet_name = get_month_sheet_name()
    print(f"\n[{now}] Syncing YOUR Jira activity → Sheet: '{sheet_name}'")

    state        = load_state()
    wb           = load_or_create_workbook()
    ws           = get_or_create_month_sheet(wb, sheet_name)
    my_id        = get_my_account_id()
    issues       = fetch_my_issues(my_id)
    total_logged = 0
    log_entries  = []

    print(f"  Found {len(issues)} ticket(s) assigned to you.")

    ticket_rows_map = read_existing_sheet_data(ws)

    for issue in issues:
        key      = issue["key"]
        fields   = issue["fields"]
        summary  = fields.get("summary", "")
        priority = (fields.get("priority") or {}).get("name", "")

        logged_ids = set(state.get(key, {}).get("logged_ids", []))
        changelog  = fetch_changelog(key)
        windows    = get_my_assignment_windows(changelog, my_id)

        if not windows:
            windows = [("1970-01-01T00:00:00.000+0000", None)]

        new_rows = []

        # ── ASSIGNMENT + STATUS CHANGES ──────────────────────
        for entry in changelog:
            entry_id   = entry["id"]
            timestamp  = entry["created"]
            updated_by = entry.get("author", {}).get("displayName", "Unknown")

            if entry_id in logged_ids:
                continue
            if not is_current_month(timestamp):
                continue

            for item in entry.get("items", []):

                # Ticket assigned to me
                if item["field"] == "assignee" and item.get("to") == my_id:
                    done_at = to_ist(timestamp)
                    status_at_assignment = "—"
                    for prev in sorted(changelog, key=lambda x: x["created"]):
                        if prev["created"] > timestamp:
                            break
                        for pi in prev.get("items", []):
                            if pi["field"] == "status":
                                status_at_assignment = pi.get("toString", "—")
                    if status_at_assignment == "—":
                        for prev in sorted(changelog, key=lambda x: x["created"]):
                            for pi in prev.get("items", []):
                                if pi["field"] == "status":
                                    status_at_assignment = pi.get("fromString", "—")
                                    break
                            if status_at_assignment != "—":
                                break

                    # ★ Done At (col 7) is before Updated By (col 8)
                    row = [key, summary, "👤 Ticket Assigned to Me", "—", status_at_assignment, "", done_at, updated_by, priority]
                    new_rows.append(row)
                    logged_ids.add(entry_id)
                    total_logged += 1
                    print(f"  ✓ {key}: assigned by {updated_by} at {done_at}")
                    log_entries.append(f"👤 {key}: Assigned by {updated_by} | Status: {status_at_assignment} | At: {done_at}")

                # ★ Only tracked status transitions
                elif item["field"] == "status" and was_assigned_to_me(timestamp, windows):
                    old_status = item.get("fromString", "—")
                    new_status = item.get("toString", "—")
                    logged_ids.add(entry_id)   # always mark as seen to avoid re-checking

                    if not is_tracked_transition(old_status, new_status):
                        continue               # skip untracked transitions silently

                    done_at = to_ist(timestamp)
                    # ★ Done At (col 7) is before Updated By (col 8)
                    row = [key, summary, "🔄 Status Changed", old_status, new_status, "", done_at, updated_by, priority]
                    new_rows.append(row)
                    total_logged += 1
                    print(f"  ✓ {key}: {old_status} → {new_status} by {updated_by} at {done_at}")
                    log_entries.append(f"🔄 {key}: {old_status} → {new_status} by {updated_by} | At: {done_at}")

        # ── COMMENTS ─────────────────────────────────────────
        comments = fetch_comments(key)
        for comment in comments:
            comment_id = comment["id"]
            timestamp  = comment["created"]
            author_id  = comment.get("author", {}).get("accountId", "")
            updated_by = comment.get("author", {}).get("displayName", "Unknown")

            if comment_id in logged_ids:
                continue
            if not is_current_month(timestamp):
                continue
            if author_id != my_id:
                continue
            if not was_assigned_to_me(timestamp, windows):
                continue

            text    = extract_comment_text(comment.get("body", {}))
            done_at = to_ist(timestamp)
            # ★ Done At (col 7) is before Updated By (col 8)
            row = [key, summary, "💬 Comment Added", "", "", text, done_at, updated_by, priority]
            new_rows.append(row)
            logged_ids.add(comment_id)
            total_logged += 1
            print(f"  ✓ {key}: comment by {updated_by} at {done_at}")
            log_entries.append(f"💬 {key}: Comment by {updated_by} | At: {done_at}")

        if new_rows:
            ticket_rows_map[key].extend(new_rows)

        if key not in state:
            state[key] = {}
        state[key]["logged_ids"] = list(logged_ids)

    rebuild_sheet(ws, ticket_rows_map)
    update_run_log(wb, total_logged, log_entries, len(issues))
    wb.save(OUTPUT_FILE)

    if total_logged:
        print(f"\n  → Saved {total_logged} new entries to '{OUTPUT_FILE}' → sheet '{sheet_name}'")
    elif not ticket_rows_map:
        print(f"\n  → No activity found. Empty sheet '{sheet_name}' created in '{OUTPUT_FILE}'")
    else:
        print("\n  → No new activity found.")

    save_state(state)

if __name__ == "__main__":
    sync()
